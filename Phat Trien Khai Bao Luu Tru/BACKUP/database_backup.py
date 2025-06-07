import sqlite3
import os
import shutil
from datetime import datetime
from typing import Optional, List, Dict, Any
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class Database:
    def __init__(self, app_dir: str):
        """Khởi tạo kết nối database"""
        self.app_dir = app_dir
        self.data_dir = os.path.join(app_dir, "data")
        self.images_dir = os.path.join(self.data_dir, "images")
        self.db_path = os.path.join(self.data_dir, "database.db")
        
        # Tạo thư mục data nếu chưa tồn tại
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        # Tạo kết nối và bảng nếu chưa tồn tại
        self.create_tables()
    
    def _get_base_dir(self) -> str:
        """Lấy thư mục gốc của ứng dụng"""
        if getattr(sys, 'frozen', False):
            # Nếu đang chạy từ file exe
            return os.path.dirname(sys.executable)
        else:
            # Nếu đang chạy từ source code
            return self.app_dir

    def create_tables(self):
        """Tạo các bảng trong database nếu chưa tồn tại"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Tạo bảng công dân - thêm id tự tăng làm khóa chính
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS cong_dan (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    so_giay_to TEXT NOT NULL,
                    so_cmnd_cu TEXT,
                    ho_ten TEXT NOT NULL,
                    gioi_tinh TEXT,
                    ngay_sinh TEXT,
                    noi_thuong_tru TEXT,
                    ngay_cap TEXT,
                    loai_giay_to TEXT,
                    ten_phong TEXT,
                    thoi_gian_ghi TEXT NOT NULL,
                    anh_mat_truoc TEXT,
                    anh_mat_sau TEXT
                )
            """)
            
            conn.commit()
        except Exception as e:
            print(f"Lỗi khi tạo bảng: {e}")
        finally:
            if conn:
                conn.close()

    def them_cong_dan(self, data: Dict[str, Any]) -> bool:
        """Thêm thông tin công dân mới vào database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Kiểm tra dữ liệu bắt buộc
            if not data["so_giay_to"] or not data["ho_ten"]:
                return False, "Số giấy tờ và họ tên là bắt buộc"
            
            # Thêm dữ liệu vào bảng
            cursor.execute("""
                INSERT INTO cong_dan (
                    so_giay_to, so_cmnd_cu, ho_ten, gioi_tinh,
                    ngay_sinh, noi_thuong_tru, ngay_cap, loai_giay_to,
                    ten_phong, thoi_gian_ghi, anh_mat_truoc, anh_mat_sau
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data["so_giay_to"], data["so_cmnd_cu"], data["ho_ten"],
                data["gioi_tinh"], data["ngay_sinh"], data["noi_thuong_tru"],
                data["ngay_cap"], data["loai_giay_to"], data["ten_phong"],
                data["thoi_gian_ghi"], data["anh_mat_truoc"], data["anh_mat_sau"]
            ))
            
            conn.commit()
            return True, "Thêm thành công"
            
        except Exception as e:
            return False, str(e)
        finally:
            if conn:
                conn.close()

    def cap_nhat_cong_dan(self, data, record_id):
        """Cập nhật thông tin công dân"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Kiểm tra dữ liệu bắt buộc
            if not data["so_giay_to"] or not data["ho_ten"]:
                return False, "Số giấy tờ và họ tên là bắt buộc"
            
            # Cập nhật thông tin
            cursor.execute("""
                UPDATE cong_dan SET
                    so_giay_to = ?,
                    so_cmnd_cu = ?,
                    ho_ten = ?,
                    gioi_tinh = ?,
                    ngay_sinh = ?,
                    noi_thuong_tru = ?,
                    ngay_cap = ?,
                    loai_giay_to = ?,
                    ten_phong = ?,
                    anh_mat_truoc = ?,
                    anh_mat_sau = ?
                WHERE id = ?
            """, (
                data["so_giay_to"], data["so_cmnd_cu"], data["ho_ten"],
                data["gioi_tinh"], data["ngay_sinh"], data["noi_thuong_tru"],
                data["ngay_cap"], data["loai_giay_to"], data["ten_phong"],
                data["anh_mat_truoc"], data["anh_mat_sau"],
                record_id
            ))
            
            if cursor.rowcount == 0:
                return False, "Không tìm thấy thông tin để cập nhật"
            
            conn.commit()
            return True, "Cập nhật thành công"
            
        except Exception as e:
            return False, str(e)
        finally:
            if conn:
                conn.close()

    def xoa_cong_dan_theo_dong(self, so_giay_to: str, thoi_gian_ghi: str) -> bool:
        """Xóa thông tin một công dân"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Lấy thông tin ảnh trước khi xóa
            cursor.execute("""
                SELECT anh_mat_truoc, anh_mat_sau
                FROM cong_dan
                WHERE so_giay_to = ? AND thoi_gian_ghi = ?
            """, (so_giay_to, thoi_gian_ghi))
            result = cursor.fetchone()
            
            if result:
                front_img, back_img = result
                # Xóa file ảnh nếu tồn tại
                for img_path in [front_img, back_img]:
                    if img_path:
                        full_path = os.path.join(self.app_dir, img_path)
                        if os.path.exists(full_path):
                            os.remove(full_path)
            
            # Xóa bản ghi
            cursor.execute("""
                DELETE FROM cong_dan
                WHERE so_giay_to = ? AND thoi_gian_ghi = ?
            """, (so_giay_to, thoi_gian_ghi))
            
            if cursor.rowcount == 0:
                return False, "Không tìm thấy thông tin để xóa"
            
            conn.commit()
            return True, "Xóa thành công"
            
        except Exception as e:
            return False, str(e)
        finally:
            if conn:
                conn.close()

    def xoa_tat_ca_cong_dan(self) -> bool:
        """Xóa tất cả thông tin công dân"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Lấy tất cả đường dẫn ảnh
            cursor.execute("SELECT anh_mat_truoc, anh_mat_sau FROM cong_dan")
            results = cursor.fetchall()
            
            # Xóa tất cả file ảnh
            for front_img, back_img in results:
                for img_path in [front_img, back_img]:
                    if img_path:
                        full_path = os.path.join(self.app_dir, img_path)
                        if os.path.exists(full_path):
                            os.remove(full_path)
            
            # Xóa tất cả bản ghi
            cursor.execute("DELETE FROM cong_dan")
            conn.commit()
            
            return True, "Đã xóa tất cả thông tin"
            
        except Exception as e:
            return False, str(e)
        finally:
            if conn:
                conn.close()

    def tim_kiem_theo_ten_va_ngay(self, search_text="", from_date=None, to_date=None, sort_order="DESC"):
        """Tìm kiếm công dân theo tên và khoảng thời gian"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Chuẩn bị câu truy vấn
            query = "SELECT * FROM cong_dan WHERE 1=1"
            params = []
            
            # Thêm điều kiện tìm kiếm theo tên
            if search_text:
                query += " AND ho_ten LIKE ?"
                params.append(f"%{search_text}%")
            
            # Thêm điều kiện ngày chỉ khi có cả from_date và to_date
            if from_date and to_date:
                query += " AND thoi_gian_ghi >= ? AND thoi_gian_ghi <= ?"
                params.extend([from_date, to_date])
            
            # Sắp xếp theo thời gian ghi
            order_by = "DESC" if sort_order.upper() == "DESC" else "ASC"
            query += f" ORDER BY thoi_gian_ghi {order_by}"
            
            print(f"Query: {query}")
            print(f"Params: {params}")
            
            # Thực thi truy vấn
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # Chuyển kết quả thành list of dict
            columns = [desc[0] for desc in cursor.description]
            data = []
            for row in results:
                data.append(dict(zip(columns, row)))
            
            print(f"Found {len(data)} records")
            return data
            
        except Exception as e:
            print(f"Lỗi khi tìm kiếm: {e}")
            return []
        finally:
            if conn:
                conn.close()

    def xuat_excel(self, excel_path, sort_order="DESC"):
        """Xuất tất cả dữ liệu ra file Excel"""
        try:
            conn = sqlite3.connect(self.db_path)
            
            # Đọc dữ liệu từ database với thứ tự sắp xếp
            order_by = "DESC" if sort_order.upper() == "DESC" else "ASC"
            query = f"SELECT * FROM cong_dan ORDER BY thoi_gian_ghi {order_by}"
            df = pd.read_sql_query(query, conn)
            
            # Đổi tên cột
            column_names = {
                'id': 'ID',
                'so_giay_to': 'Số giấy tờ',
                'so_cmnd_cu': 'Số CMND cũ',
                'ho_ten': 'Họ và tên',
                'gioi_tinh': 'Giới tính',
                'ngay_sinh': 'Ngày sinh',
                'noi_thuong_tru': 'Nơi thường trú',
                'ngay_cap': 'Ngày cấp',
                'loai_giay_to': 'Loại giấy tờ',
                'ten_phong': 'Tên phòng',
                'thoi_gian_ghi': 'Thời gian ghi',
                'anh_mat_truoc': 'Ảnh mặt trước',
                'anh_mat_sau': 'Ảnh mặt sau'
            }
            df = df.rename(columns=column_names)
            
            # Tạo workbook mới
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Thiết lập font mặc định cho toàn bộ sheet
            ws.font = Font(name='Times New Roman', size=11)
            
            # Ghi header
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(name='Times New Roman', size=12, bold=True)
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Ghi dữ liệu
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Xử lý đặc biệt cho cột ảnh
                    if col_idx in [12, 13]:  # Cột ảnh mặt trước và sau
                        if value and str(value).strip():
                            # Tạo đường dẫn tuyệt đối từ thư mục Excel đến file ảnh
                            img_path = os.path.abspath(os.path.join(self.app_dir, value))
                            cell.value = f'=HYPERLINK("{img_path}", "Xem ảnh")'
                            cell.font = Font(name='Times New Roman', size=11, color="0000FF", underline="single")
                        else:
                            cell.value = "Không có ảnh"
                            cell.font = Font(name='Times New Roman', size=11)
                    else:
                        cell.value = value
                        cell.font = Font(name='Times New Roman', size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo border style
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Áp dụng border cho tất cả các ô có dữ liệu
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            # Tự động điều chỉnh độ rộng cột dựa trên nội dung
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Thêm hệ số 1.2 để có thêm khoảng trống
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Thiết lập chiều cao tối thiểu cho các dòng
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 25
            
            # Lưu file
            wb.save(excel_path)
            
            return True
            
        except Exception as e:
            print(f"Lỗi khi xuất Excel: {e}")
            return False
        finally:
            if conn:
                conn.close()

    def xuat_excel_tu_ket_qua(self, excel_path, data, sort_order="DESC"):
        """Xuất kết quả tìm kiếm ra file Excel"""
        try:
            # Tạo DataFrame từ dữ liệu
            df = pd.DataFrame(data)
            
            # Sắp xếp dữ liệu theo thời gian ghi
            df = df.sort_values(by='thoi_gian_ghi', 
                              ascending=False if sort_order.upper() == "DESC" else True)
            
            # Đổi tên cột
            column_names = {
                'id': 'ID',
                'so_giay_to': 'Số giấy tờ',
                'so_cmnd_cu': 'Số CMND cũ',
                'ho_ten': 'Họ và tên',
                'gioi_tinh': 'Giới tính',
                'ngay_sinh': 'Ngày sinh',
                'noi_thuong_tru': 'Nơi thường trú',
                'ngay_cap': 'Ngày cấp',
                'loai_giay_to': 'Loại giấy tờ',
                'ten_phong': 'Tên phòng',
                'thoi_gian_ghi': 'Thời gian ghi',
                'anh_mat_truoc': 'Ảnh mặt trước',
                'anh_mat_sau': 'Ảnh mặt sau'
            }
            df = df.rename(columns=column_names)
            
            # Tạo workbook mới
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Thiết lập font mặc định cho toàn bộ sheet
            ws.font = Font(name='Times New Roman', size=11)
            
            # Ghi header
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(name='Times New Roman', size=12, bold=True)
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Ghi dữ liệu
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Xử lý đặc biệt cho cột ảnh
                    if col_idx in [12, 13]:  # Cột ảnh mặt trước và sau
                        if value and str(value).strip():
                            # Tạo đường dẫn tuyệt đối từ thư mục Excel đến file ảnh
                            img_path = os.path.abspath(os.path.join(self.app_dir, value))
                            cell.value = f'=HYPERLINK("{img_path}", "Xem ảnh")'
                            cell.font = Font(name='Times New Roman', size=11, color="0000FF", underline="single")
                        else:
                            cell.value = "Không có ảnh"
                            cell.font = Font(name='Times New Roman', size=11)
                    else:
                        cell.value = value
                        cell.font = Font(name='Times New Roman', size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo border style
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Áp dụng border cho tất cả các ô có dữ liệu
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            # Tự động điều chỉnh độ rộng cột dựa trên nội dung
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Thêm hệ số 1.2 để có thêm khoảng trống
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Thiết lập chiều cao tối thiểu cho các dòng
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 25
            
            # Lưu file
            wb.save(excel_path)
            
            return True
            
        except Exception as e:
            print(f"Lỗi khi xuất Excel: {e}")
            return False

    def luu_anh(self, source_path: str, prefix: str = "") -> str:
        """Lưu ảnh vào thư mục images"""
        if not source_path:
            return ""
            
        # Tạo tên file mới
        ext = os.path.splitext(source_path)[1]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"{prefix}_{timestamp}{ext}"
        new_path = os.path.join("data", "images", new_filename)
        
        # Copy file ảnh
        dest_path = os.path.join(self.images_dir, new_filename)
        shutil.copy2(source_path, dest_path)
        
        return new_path 