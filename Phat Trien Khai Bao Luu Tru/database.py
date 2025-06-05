import sqlite3
import os
import shutil
from datetime import datetime
from typing import Optional, List, Dict, Any
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class Database:
    def __init__(self, app_dir: str):
        # Tạo cấu trúc thư mục
        self.app_dir = app_dir
        self.data_dir = os.path.join(app_dir, "data")
        self.images_dir = os.path.join(self.data_dir, "images")
        self.db_path = os.path.join(self.data_dir, "cong_dan.db")
        
        # Đảm bảo các thư mục tồn tại
        os.makedirs(self.data_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)
        
        self._create_tables()
    
    def _get_base_dir(self) -> str:
        """Lấy thư mục gốc của ứng dụng"""
        if getattr(sys, 'frozen', False):
            # Nếu đang chạy từ file exe
            return os.path.dirname(sys.executable)
        else:
            # Nếu đang chạy từ source code
            return self.app_dir

    def _create_tables(self):
        """Tạo các bảng nếu chưa tồn tại"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Bảng thông tin công dân
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS cong_dan (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ho_ten TEXT NOT NULL,
                    ngay_sinh TEXT,
                    gioi_tinh TEXT,
                    quoc_tich TEXT,
                    cmnd TEXT,
                    ho_chieu TEXT,
                    thuong_tru TEXT,
                    tam_tru TEXT,
                    nghe_nghiep TEXT,
                    email TEXT,
                    sdt TEXT,
                    anh_mat_truoc TEXT,
                    anh_mat_sau TEXT,
                    ngay_den TEXT,
                    ngay_di TEXT,
                    phong TEXT,
                    ghi_chu TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()

    def them_cong_dan(self, data: Dict[str, Any]) -> int:
        """Thêm thông tin công dân mới"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Chuẩn bị câu lệnh SQL
            columns = ', '.join(data.keys())
            placeholders = ', '.join(['?' for _ in data])
            sql = f'INSERT INTO cong_dan ({columns}) VALUES ({placeholders})'
            
            # Thực thi câu lệnh
            cursor.execute(sql, list(data.values()))
            conn.commit()
            
            return cursor.lastrowid

    def cap_nhat_cong_dan(self, id: int, data: Dict[str, Any]) -> bool:
        """Cập nhật thông tin công dân"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Chuẩn bị câu lệnh SQL
            set_clause = ', '.join([f'{k} = ?' for k in data.keys()])
            sql = f'UPDATE cong_dan SET {set_clause} WHERE id = ?'
            
            # Thực thi câu lệnh
            values = list(data.values()) + [id]
            cursor.execute(sql, values)
            conn.commit()
            
            return cursor.rowcount > 0

    def xoa_cong_dan(self, id: int) -> bool:
        """Xóa thông tin công dân"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Lấy thông tin ảnh trước khi xóa
            cursor.execute('SELECT anh_mat_truoc, anh_mat_sau FROM cong_dan WHERE id = ?', (id,))
            result = cursor.fetchone()
            if result:
                anh_mat_truoc, anh_mat_sau = result
                
                # Xóa file ảnh nếu tồn tại
                if anh_mat_truoc:
                    try:
                        os.remove(os.path.join(self.images_dir, os.path.basename(anh_mat_truoc)))
                    except:
                        pass
                if anh_mat_sau:
                    try:
                        os.remove(os.path.join(self.images_dir, os.path.basename(anh_mat_sau)))
                    except:
                        pass
            
            # Xóa record trong database
            cursor.execute('DELETE FROM cong_dan WHERE id = ?', (id,))
            conn.commit()
            
            return cursor.rowcount > 0

    def lay_cong_dan(self, id: int) -> Optional[Dict[str, Any]]:
        """Lấy thông tin công dân theo ID"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM cong_dan WHERE id = ?', (id,))
            columns = [col[0] for col in cursor.description]
            result = cursor.fetchone()
            
            if result:
                return dict(zip(columns, result))
            return None

    def lay_danh_sach_cong_dan(self) -> List[Dict[str, Any]]:
        """Lấy danh sách tất cả công dân"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM cong_dan ORDER BY id')
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
            
            return [dict(zip(columns, row)) for row in results]

    def xuat_excel(self) -> str:
        """Xuất danh sách công dân ra file Excel"""
        # Lấy danh sách công dân
        data = self.lay_danh_sach_cong_dan()
        
        # Tạo DataFrame
        df = pd.DataFrame(data)
        
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách công dân"
        
        # Định dạng header
        headers = [
            'ID', 'Họ tên', 'Ngày sinh', 'Giới tính', 'Quốc tịch', 'CMND/CCCD', 'Hộ chiếu',
            'Thường trú', 'Tạm trú', 'Nghề nghiệp', 'Email', 'Số điện thoại',
            'Ảnh mặt trước', 'Ảnh mặt sau', 'Ngày đến', 'Ngày đi', 'Phòng', 'Ghi chú', 'Ngày tạo'
        ]
        
        # Thêm header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thêm dữ liệu
        for row_idx, item in enumerate(data, 2):
            # ID
            ws.cell(row=row_idx, column=1, value=item['id'])
            
            # Các trường thông tin cơ bản
            ws.cell(row=row_idx, column=2, value=item['ho_ten'])
            ws.cell(row=row_idx, column=3, value=item['ngay_sinh'])
            ws.cell(row=row_idx, column=4, value=item['gioi_tinh'])
            ws.cell(row=row_idx, column=5, value=item['quoc_tich'])
            ws.cell(row=row_idx, column=6, value=item['cmnd'])
            ws.cell(row=row_idx, column=7, value=item['ho_chieu'])
            ws.cell(row=row_idx, column=8, value=item['thuong_tru'])
            ws.cell(row=row_idx, column=9, value=item['tam_tru'])
            ws.cell(row=row_idx, column=10, value=item['nghe_nghiep'])
            ws.cell(row=row_idx, column=11, value=item['email'])
            ws.cell(row=row_idx, column=12, value=item['sdt'])
            
            # Xử lý đường dẫn ảnh
            base_dir = self._get_base_dir()
            
            # Ảnh mặt trước
            if item['anh_mat_truoc']:
                image_path = os.path.abspath(os.path.join(base_dir, item['anh_mat_truoc']))
                url_path = f"file:///{image_path.replace(os.sep, '/')}"
                ws.cell(row=row_idx, column=13, value=f'=HYPERLINK("{url_path}","Xem ảnh")')
            
            # Ảnh mặt sau
            if item['anh_mat_sau']:
                image_path = os.path.abspath(os.path.join(base_dir, item['anh_mat_sau']))
                url_path = f"file:///{image_path.replace(os.sep, '/')}"
                ws.cell(row=row_idx, column=14, value=f'=HYPERLINK("{url_path}","Xem ảnh")')
            
            # Các trường thông tin khác
            ws.cell(row=row_idx, column=15, value=item['ngay_den'])
            ws.cell(row=row_idx, column=16, value=item['ngay_di'])
            ws.cell(row=row_idx, column=17, value=item['phong'])
            ws.cell(row=row_idx, column=18, value=item['ghi_chu'])
            ws.cell(row=row_idx, column=19, value=item['created_at'])
        
        # Căn chỉnh độ rộng cột
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Tạo tên file với timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(self.data_dir, f"DS_Cong_Dan_{timestamp}.xlsx")
        
        # Lưu file
        wb.save(excel_path)
        
        return excel_path

    def luu_anh(self, source_path: str, prefix: str = "") -> str:
        """Lưu ảnh vào thư mục images"""
        if not source_path:
            return ""
            
        # Tạo tên file mới
        ext = os.path.splitext(source_path)[1]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"{prefix}_{timestamp}{ext}"
        new_path = os.path.join("data", "images", new_filename)
        
        # Copy file ảnh
        dest_path = os.path.join(self.images_dir, new_filename)
        shutil.copy2(source_path, dest_path)
        
        return new_path 